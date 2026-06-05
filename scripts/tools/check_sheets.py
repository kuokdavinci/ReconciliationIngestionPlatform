import openpyxl

wb = openpyxl.load_workbook("/home/kuokdavinci/AdapterService/RequestTemplate.xlsx")
print("Sheets in RequestTemplate.xlsx:", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(f"Sheet {name}: Max Row = {ws.max_row}, Max Col = {ws.max_column}")
