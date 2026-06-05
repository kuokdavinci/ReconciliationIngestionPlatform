import openpyxl
import sys
from pathlib import Path

def inspect_xlsx(path):
    print(f"=== Inspecting {path} ===")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    print(f"Sheet Name: {ws.title}")
    print(f"Max Row: {ws.max_row}, Max Col: {ws.max_column}")
    
    # Print the first 15 rows to see structure
    for r in range(1, min(20, ws.max_row + 1)):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        # Filter empty values at the end of the row representation for cleaner print
        while row_vals and row_vals[-1] is None:
            row_vals.pop()
        if any(v is not None for v in row_vals):
            print(f"Row {r}: {row_vals}")
        else:
            print(f"Row {r}: [ALL EMPTY]")

print("INSPECTING ORIGINALS")
inspect_xlsx("/home/kuokdavinci/AdapterService/RequestTemplate.xlsx")
inspect_xlsx("/home/kuokdavinci/AdapterService/sftp_data/m4becomvsp_07072024_combine.xlsx")
