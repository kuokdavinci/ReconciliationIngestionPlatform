"""Structure signature computation for file format detection.

Computes a fingerprint of a data file's structure (headers, column count,
data type patterns) independent of any MappingConfig. Used by ConfigHealthService
to detect when a partner's file format has changed.
"""

import csv
import hashlib
import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any


@dataclass
class StructureSignature:
    """Fingerprint of a data file's columnar structure.

    Attributes:
        headers: First data row (typically headers/column names).
        column_count: Number of columns in the data.
        sample_rows: Up to 10 sample data rows for AI analysis.
        hash: MD5 of headers + column_count for quick staleness check.
    """

    headers: list[str] = field(default_factory=list)
    column_count: int = 0
    sample_rows: list[list[str]] = field(default_factory=list)
    hash: str = ""
    header_row_index: int = 1
    first_data_row_index: int = 2

    def to_dict(self) -> dict[str, Any]:
        return {
            "headers": self.headers,
            "columnCount": self.column_count,
            "sampleRows": self.sample_rows[:5],
            "hash": self.hash,
            "headerRowIndex": self.header_row_index,
            "firstDataRowIndex": self.first_data_row_index,
        }

    @classmethod
    def from_dict(cls, data: dict[str, Any]) -> "StructureSignature":
        return cls(
            headers=data.get("headers", []),
            column_count=data.get("columnCount", 0),
            sample_rows=data.get("sampleRows", []),
            hash=data.get("hash", ""),
            header_row_index=data.get("headerRowIndex", 1),
            first_data_row_index=data.get("firstDataRowIndex", 2),
        )


def _normalize_cell(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _compute_hash(headers: list[str], column_count: int) -> str:
    raw = json.dumps([headers, column_count], sort_keys=True, ensure_ascii=False)
    return hashlib.md5(raw.encode()).hexdigest()


def structure_signature_shape(signature: Any) -> tuple[tuple[str, ...], int] | None:
    """Return the stable shape portion of a persisted or computed signature."""
    if isinstance(signature, StructureSignature):
        headers = signature.headers
        column_count = signature.column_count
    elif isinstance(signature, dict):
        headers = signature.get("headers")
        column_count = signature.get("columnCount")
    else:
        return None

    if not isinstance(headers, list) or not headers:
        return None
    if not isinstance(column_count, int):
        column_count = len(headers)
    return tuple(str(header) for header in headers), column_count


def structure_signatures_equivalent(left: Any, right: Any) -> bool:
    """Compare file structure without considering sample row values or hashes."""
    left_shape = structure_signature_shape(left)
    right_shape = structure_signature_shape(right)
    return left_shape is not None and left_shape == right_shape


def _read_raw_csv(path: Path, max_rows: int = 20) -> list[list[str]]:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f, delimiter=delimiter)
        rows = []
        for row in reader:
            normalized = [_normalize_cell(c) for c in row]
            if not rows and not any(normalized):
                continue
            if len(rows) >= max_rows:
                break
            rows.append(normalized)
    return rows


def _read_raw_xlsx(path: Path, max_rows: int = 20) -> list[list[str]]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        normalized = [_normalize_cell(c) for c in row]
        if not rows and not any(normalized):
            continue
        if len(rows) >= max_rows:
            break
        rows.append(normalized)
    wb.close()
    return rows


def _read_raw_xlsx_with_indices(
    path: Path,
    max_rows: int = 20,
) -> list[tuple[int, list[str]]]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    ws = wb.active
    rows: list[tuple[int, list[str]]] = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        normalized = [_normalize_cell(c) for c in row]
        if not rows and not any(normalized):
            continue
        if len(rows) >= max_rows:
            break
        rows.append((row_idx, normalized))
    wb.close()
    return rows


def _read_raw_json(path: Path, max_rows: int = 20) -> list[list[str]]:
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    rows: list[list[str]] = []
    if isinstance(data, dict) and isinstance(data.get("items"), list):
        data = data["items"]
    if not isinstance(data, list):
        raise ValueError("JSON root must be an array or an object with an items array")
    for i, item in enumerate(data):
        if i >= max_rows:
            break
        if isinstance(item, list):
            rows.append([_normalize_cell(c) for c in item])
        elif isinstance(item, dict):
            rows.append([_normalize_cell(v) for v in item.values()])
        else:
            rows.append([_normalize_cell(item)])
    return rows


def read_raw_rows(path: str | Path, max_rows: int = 20) -> list[list[str]]:
    """Read raw rows from any supported data file without needing a MappingConfig.

    Used by compute_signature() and AIConfigGenerator to inspect file structure.
    """
    p = Path(path)
    suffix = p.suffix.lower()
    if suffix == ".json":
        return _read_raw_json(p, max_rows)
    if suffix in (".xlsx", ".xlsm"):
        return _read_raw_xlsx(p, max_rows)
    if suffix in (".csv", ".tsv"):
        return _read_raw_csv(p, max_rows)
    raise ValueError(f"Unsupported file extension: {suffix}")


def compute_signature(file_path: str | Path, sample_size: int = 10) -> StructureSignature:
    """Compute a StructureSignature from a data file.

    Process:
    1. Read first (sample_size + 1) raw rows.
    2. Treat first row as headers.
    3. Compute column count and MD5 hash.
    4. Return StructureSignature with up to sample_size data rows.
    """
    p = Path(file_path)
    if p.suffix.lower() == ".json":
        with open(p, encoding="utf-8") as f:
            data = json.load(f)
        items = data.get("items") if isinstance(data, dict) else data
        if isinstance(items, list) and items and all(isinstance(item, dict) for item in items):
            headers: list[str] = []
            for item in items[: sample_size + 1]:
                for key in item:
                    if key not in headers:
                        headers.append(str(key))
            sample = [
                [_normalize_cell(item.get(header)) for header in headers]
                for item in items[:sample_size]
            ]
            return StructureSignature(
                headers=headers,
                column_count=len(headers),
                sample_rows=sample,
                hash=_compute_hash(headers, len(headers)),
                header_row_index=1,
                first_data_row_index=1,
            )
    if p.suffix.lower() in (".xlsx", ".xlsm"):
        raw_with_indices = _read_raw_xlsx_with_indices(p, max_rows=sample_size + 1)
        if not raw_with_indices:
            return StructureSignature()
        (header_row_index, headers), *sample_indexed = raw_with_indices
        sample = [row for _, row in sample_indexed[:sample_size]]
        first_data_row_index = sample_indexed[0][0] if sample_indexed else header_row_index + 1
        column_count = len(headers)
        return StructureSignature(
            headers=headers,
            column_count=column_count,
            sample_rows=sample,
            hash=_compute_hash(headers, column_count),
            header_row_index=header_row_index,
            first_data_row_index=first_data_row_index,
        )

    raw = read_raw_rows(file_path, max_rows=sample_size + 1)
    if not raw:
        return StructureSignature()

    headers = raw[0] if len(raw) > 0 else []
    sample = raw[1: sample_size + 1] if len(raw) > 1 else []
    column_count = len(headers)

    return StructureSignature(
        headers=headers,
        column_count=column_count,
        sample_rows=sample,
        hash=_compute_hash(headers, column_count),
        header_row_index=1,
        first_data_row_index=2 if sample else 1,
    )
