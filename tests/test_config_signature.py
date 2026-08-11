"""Tests for StructureSignature computation and raw row reading."""

import tempfile
from pathlib import Path

import pytest

from src.config.signature import (
    StructureSignature,
    compute_signature,
    read_raw_rows,
)


def _path(suffix: str) -> Path:
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as f:
        return Path(f.name)


class TestReadRawRows:
    def test_csv_file(self) -> None:
        p = _path(".csv")
        p.write_text("a,b,c\n1,2,3\n4,5,6\n", encoding="utf-8")
        rows = read_raw_rows(p)
        p.unlink(missing_ok=True)
        assert rows == [["a", "b", "c"], ["1", "2", "3"], ["4", "5", "6"]]

    def test_tsv_file(self) -> None:
        p = _path(".tsv")
        p.write_text("a\tb\tc\n1\t2\t3\n", encoding="utf-8")
        rows = read_raw_rows(p)
        p.unlink(missing_ok=True)
        assert rows == [["a", "b", "c"], ["1", "2", "3"]]

    def test_xlsx_file(self) -> None:
        from openpyxl import Workbook
        p = _path(".xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["x", "y"])
        ws.append(["10", "20"])
        wb.save(str(p))
        wb.close()
        rows = read_raw_rows(p)
        p.unlink(missing_ok=True)
        assert rows == [["x", "y"], ["10", "20"]]

    def test_xlsx_skips_leading_empty_rows(self) -> None:
        from openpyxl import Workbook
        p = _path(".xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append([None, None])
        ws.append([None, None])
        ws.append(["h1", "h2"])
        ws.append(["v1", "v2"])
        wb.save(str(p))
        wb.close()
        rows = read_raw_rows(p)
        p.unlink(missing_ok=True)
        assert rows == [["h1", "h2"], ["v1", "v2"]]

    def test_json_file(self) -> None:
        import json
        p = _path(".json")
        data = [["a", "b"], ["1", "2"]]
        with open(p, "w") as f:
            json.dump(data, f)
        rows = read_raw_rows(p)
        p.unlink(missing_ok=True)
        assert rows == [["a", "b"], ["1", "2"]]

    def test_json_pagination_envelope(self) -> None:
        import json
        p = _path(".json")
        data = {
            "items": [
                {"id": "VTP-001", "amount": 100},
                {"id": "VTP-002", "amount": 200},
            ],
            "nextCursor": "cursor-1",
        }
        with open(p, "w") as f:
            json.dump(data, f)
        rows = read_raw_rows(p)
        p.unlink(missing_ok=True)
        assert rows == [["VTP-001", "100"], ["VTP-002", "200"]]

    def test_json_pagination_pages_have_stable_structure_rows(self) -> None:
        import json

        first = _path(".json")
        second = _path(".json")
        first.write_text(
            json.dumps({"items": [{"id": "VTP-001", "amount": 100, "status": "OK"}]}),
            encoding="utf-8",
        )
        second.write_text(
            json.dumps({"items": [{"id": "VTP-003", "amount": 300, "status": "OK"}]}),
            encoding="utf-8",
        )

        first_signature = compute_signature(first)
        second_signature = compute_signature(second)

        first.unlink(missing_ok=True)
        second.unlink(missing_ok=True)
        assert first_signature.headers == ["id", "amount", "status"]
        assert first_signature.hash == second_signature.hash
        assert first_signature.sample_rows == [["VTP-001", "100", "OK"]]

    def test_unsupported_extension_raises(self) -> None:
        p = _path(".txt")
        p.write_text("a\nb\n", encoding="utf-8")
        with pytest.raises(ValueError, match="Unsupported file extension"):
            read_raw_rows(p)
        p.unlink(missing_ok=True)

    def test_respects_max_rows(self) -> None:
        p = _path(".csv")
        p.write_text("\n".join(f"row{i}" for i in range(100)), encoding="utf-8")
        rows = read_raw_rows(p, max_rows=5)
        p.unlink(missing_ok=True)
        assert len(rows) == 5


class TestComputeSignature:
    def test_returns_structure_signature(self) -> None:
        p = _path(".csv")
        p.write_text("id,amount,status\n1,100,OK\n2,200,FAIL\n", encoding="utf-8")
        sig = compute_signature(p)
        p.unlink(missing_ok=True)
        assert isinstance(sig, StructureSignature)
        assert sig.column_count == 3
        assert sig.headers == ["id", "amount", "status"]
        assert len(sig.sample_rows) == 2
        assert sig.header_row_index == 1
        assert sig.first_data_row_index == 2

    def test_xlsx_signature_keeps_absolute_row_positions(self) -> None:
        from openpyxl import Workbook
        p = _path(".xlsx")
        wb = Workbook()
        ws = wb.active
        for _ in range(6):
            ws.append([None, None, None])
        ws.append(["STT", "msTransId", "msNgayHoanThanh"])
        ws.append(["1", "MOMO_TXN_9000", "2026-06-05 12:00:00"])
        wb.save(str(p))
        wb.close()
        sig = compute_signature(p)
        p.unlink(missing_ok=True)
        assert sig.header_row_index == 7
        assert sig.first_data_row_index == 8
        assert sig.headers[1] == "msTransId"

    def test_hash_consistency(self) -> None:
        p1 = _path(".csv")
        p1.write_text("a,b\n1,2\n3,4\n", encoding="utf-8")
        p2 = _path(".csv")
        p2.write_text("a,b\n1,2\n3,4\n", encoding="utf-8")
        sig1 = compute_signature(p1)
        sig2 = compute_signature(p2)
        p1.unlink(missing_ok=True)
        p2.unlink(missing_ok=True)
        assert sig1.hash == sig2.hash

    def test_hash_different_headers(self) -> None:
        p1 = _path(".csv")
        p1.write_text("a,b\n1,2\n", encoding="utf-8")
        p2 = _path(".csv")
        p2.write_text("x,y\n1,2\n", encoding="utf-8")
        sig1 = compute_signature(p1)
        sig2 = compute_signature(p2)
        p1.unlink(missing_ok=True)
        p2.unlink(missing_ok=True)
        assert sig1.hash != sig2.hash

    def test_hash_different_column_count(self) -> None:
        p1 = _path(".csv")
        p1.write_text("a,b\n1,2\n", encoding="utf-8")
        p2 = _path(".csv")
        p2.write_text("a,b,c\n1,2,3\n", encoding="utf-8")
        sig1 = compute_signature(p1)
        sig2 = compute_signature(p2)
        p1.unlink(missing_ok=True)
        p2.unlink(missing_ok=True)
        assert sig1.hash != sig2.hash

    def test_empty_file(self) -> None:
        p = _path(".csv")
        p.write_text("", encoding="utf-8")
        sig = compute_signature(p)
        p.unlink(missing_ok=True)
        assert sig.column_count == 0
        assert sig.headers == []
        assert sig.sample_rows == []


class TestStructureSignature:
    def test_to_dict(self) -> None:
        sig = StructureSignature(
            headers=["a", "b"],
            column_count=2,
            sample_rows=[["1", "2"]],
            hash="abc123",
        )
        d = sig.to_dict()
        assert d["headers"] == ["a", "b"]
        assert d["columnCount"] == 2
        assert d["hash"] == "abc123"

    def test_from_dict(self) -> None:
        d = {
            "headers": ["x", "y"],
            "columnCount": 2,
            "sampleRows": [["10"]],
            "hash": "def456",
        }
        sig = StructureSignature.from_dict(d)
        assert sig.headers == ["x", "y"]
        assert sig.column_count == 2
        assert sig.hash == "def456"
